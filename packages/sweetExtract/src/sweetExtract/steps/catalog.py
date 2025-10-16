# sweetExtract/steps/catalog.py
from __future__ import annotations
from pathlib import Path
from typing import Any, Dict, List, Optional

from sweetExtract.steps.base import BaseStep
from sweetExtract.project import Project
from sweetExtract.steps.unpack_data import BuildUnpackedData  # ensure upstream order

# minimal, dependency-light inspection
TEXT_TABLE_EXTS = {".csv", ".tsv", ".tab", ".txt"}
EXCEL_EXTS      = {".xlsx", ".xls"}
JSON_EXTS       = {".json", ".ndjson", ".jsonl"}

MAX_FILES   = 5000          # hard cap on catalog size
MAX_BYTES   = 64_000        # sample/read cap per file
MAX_SHEETS  = 24            # cap sheet introspection
MAX_COLS    = 200           # cap header width

def _has_any_files(root: Path) -> bool:
    if not root.exists():
        return False
    for p in root.rglob("*"):
        if p.is_file() and not p.name.startswith("."):
            return True
    return False

def _detect_kind(p: Path) -> str:
    ext = p.suffix.lower()
    if ext in TEXT_TABLE_EXTS: return "table/text"
    if ext in EXCEL_EXTS:      return "table/excel"
    if ext in JSON_EXTS:       return "json"
    return "other"

def _peek_text(p: Path, n: int = MAX_BYTES) -> str:
    try:
        return p.open("r", encoding="utf-8", errors="ignore").read(n)
    except Exception:
        return ""

def _sniff_delimiter(sample: str) -> Optional[str]:
    # tabs dominate if present; else choose most frequent of comma/semicolon/pipe
    counts = {
        "\t": sample.count("\t"),
        ",": sample.count(","),
        ";": sample.count(";"),
        "|": sample.count("|"),
    }
    delim, cnt = max(counts.items(), key=lambda kv: kv[1])
    return delim if cnt > 0 else None

def _first_header(sample: str, fallback_delim: str = ",") -> List[str]:
    lines = [ln for ln in sample.splitlines() if ln.strip()]
    if not lines:
        return []
    delim = _sniff_delimiter(sample) or fallback_delim
    cols = [c.strip() for c in lines[0].split(delim)]
    if len(cols) == 1 and len(lines[0]) > 200:
        # likely not a delimited header
        return []
    return cols[:MAX_COLS]

def _json_top_keys(sample: str) -> List[str]:
    # light heuristic for JSON/JSONL/NDJSON
    s = sample.lstrip()
    keys: List[str] = []
    try:
        import json as _json
        if s.startswith("{"):
            obj = _json.loads(s)
            if isinstance(obj, dict):
                keys = list(obj.keys())[:MAX_COLS]
        elif s.startswith("["):
            arr = _json.loads(s)
            if isinstance(arr, list) and arr and isinstance(arr[0], dict):
                keys = list(arr[0].keys())[:MAX_COLS]
        else:
            # assume jsonl/ndjson by lines
            for ln in s.splitlines()[:50]:
                ln = ln.strip()
                if not ln:
                    continue
                try:
                    obj = _json.loads(ln)
                except Exception:
                    continue
                if isinstance(obj, dict):
                    keys = list(obj.keys())[:MAX_COLS]
                    break
    except Exception:
        pass
    return keys

def _excel_sheet_headers(p: Path) -> Dict[str, List[str]]:
    headers: Dict[str, List[str]] = {}
    try:
        if p.suffix.lower() == ".xlsx":
            import openpyxl  # type: ignore
            wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
            for name in wb.sheetnames[:MAX_SHEETS]:
                ws = wb[name]
                row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
                if row:
                    headers[name] = [("" if v is None else str(v).strip()) for v in row][:MAX_COLS]
                else:
                    headers[name] = []
        elif p.suffix.lower() == ".xls":
            import xlrd  # type: ignore
            wb = xlrd.open_workbook(p)
            for i in range(min(wb.nsheets, MAX_SHEETS)):
                sh = wb.sheet_by_index(i)
                if sh.nrows > 0:
                    row = sh.row_values(0)
                    headers[sh.name] = [str(v).strip() for v in row][:MAX_COLS]
                else:
                    headers[sh.name] = []
    except Exception:
        headers = {}
    return headers

class Catalog(BaseStep):
    """
    Build a compact catalog of files under artifacts/data_unpacked/.
    Minimal, LLM-friendly facts only (kind, size, headers/keys, loader hints).

    Artifact: artifacts/meta/catalog.json
    """

    def __init__(self, force: bool = False):
        super().__init__(
            name="catalog",
            artifact="meta/catalog.json",
            depends_on=[BuildUnpackedData],   # ensure unpacked data exists
            map_over=None,
        )
        self._force = bool(force)

    def should_run(self, project: Project) -> bool:
        if self._force:
            return True
        root = (project.artifacts_dir / "data_unpacked").resolve()
        out  = (project.artifacts_dir / self.artifact).resolve()
        return _has_any_files(root) and not out.exists()

    def compute(self, project: Project) -> Dict[str, Any]:
        art_root = project.artifacts_dir.resolve()
        base = (art_root / "data_unpacked").resolve()

        files: List[Dict[str, Any]] = []
        total = 0

        for p in sorted(base.rglob("*")):
            if not p.is_file():
                continue
            if p.name.startswith(".") or p.name in {".DS_Store", ".gitkeep"}:
                continue

            kind = _detect_kind(p)
            rec: Dict[str, Any] = {
                "relpath": str(p.relative_to(base)),
                "size_bytes": int(p.stat().st_size),
                "ext": p.suffix.lower(),
                "kind": kind,  # "table/text" | "table/excel" | "json" | "other"
            }

            if kind == "table/text":
                sample = _peek_text(p, MAX_BYTES)
                rec["delimiter"] = _sniff_delimiter(sample) or ","
                rec["columns_head"] = _first_header(sample, rec["delimiter"])
                rec["suggested_loader"] = "pandas_read_csv"

            elif kind == "table/excel":
                rec["sheet_headers"] = _excel_sheet_headers(p)  # may be {}
                rec["suggested_loader"] = "pandas_read_excel"

            elif kind == "json":
                sample = _peek_text(p, MAX_BYTES)
                rec["top_level_keys_sample"] = _json_top_keys(sample)
                rec["suggested_loader"] = (
                    "pandas_read_json_lines" if rec["ext"] in {".ndjson", ".jsonl"} else "json_load"
                )

            else:
                rec["suggested_loader"] = "unknown"

            files.append(rec)
            total += 1
            if total >= MAX_FILES:
                break

        # summary
        by_kind: Dict[str, int] = {}
        by_ext: Dict[str, int] = {}
        for r in files:
            by_kind[r["kind"]] = by_kind.get(r["kind"], 0) + 1
            by_ext[r["ext"]]   = by_ext.get(r["ext"], 0) + 1

        return {
            "root": str(base),
            "summary": {
                "n_files": len(files),
                "by_kind": by_kind,
                "by_ext": by_ext,
            },
            "files": files,
            "notes": [
                "Only sampled bytes are read; large files are not fully parsed.",
                "Excel headers are included only if openpyxl/xlrd are available; otherwise sheet list may be empty.",
            ],
        }
